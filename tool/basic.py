import pymssql
from prettytable import PrettyTable
import os
import xlwt,xlrd # excel 写入
import openpyxl # excel 2007

# 处理pyinstaller 和 pymssql 闪退问题
import uuid
import _mssql
import decimal
import pypyodbc
decimal.__version__
uuid.ctypes.__version__
_mssql.__version__

# 处理pyinstaller 和 pymssql 闪退问题

##############################################公共工具函数#####################################################################
############################################## 数据库类 start ##############################################
class db:
	def __init__(self,host,username,password,database):
		self.host = host
		self.username = username
		self.password = password
		self.database = database
		(self.conn,self.cur) = self.conndb()
	
	# 链接数据库
	def conndb(self):
		conn = pymssql.connect(host=self.host,user=self.username,password=self.password,database=self.database,charset='utf8')
		cur = conn.cursor()
		return (conn,cur)

	# 关闭数据库链接
	def connClose(self):
		self.conn.close()

	# 数据库查询操作
	def select(self,sql):
		self.cur.execute(sql)
		data = self.cur.fetchall()
		if len(data)>0:
			return data
		else:
			return None
	
	# 数据库增删改操作
	def commit(self,sql):
		self.cur.execute(sql)
		self.conn.commit()

	#批量插入
	def insert(self,sql,data):
		self.cur.executemany(sql,data)
		self.conn.commit()
############################################## 数据库类 end ##############################################

# 格式化 输出
def formatTable(data,title,setting={}):
	x = PrettyTable(title)
	x.padding_width = 1
	if len(setting)>0:
		for key in sorted(setting):	
			if setting[key] == 'left':
				x.align[key] = 'l'
	if len(title) == len(data[0]):
		for line in data:
			x.add_row(line)
	else:
		for line in data:
			lineData = []
			for index in range(0,int(len(title))-1):
				lineData.append(line[index])
			x.add_row(lineData)
	print(x)

# 递归
def callbackFunc(message,func):
	check = input(message)
	if check == 'y':
		func()

# 遍历文件目录
def listDri(url,setting={}):
	files = os.listdir(url)
	if not ('print' in setting and setting['print'] == 'off'):
		for index in range(0,len(files)):
			print('%d.%s' % (index+1,files[index]))
	return files

# excel 操作类	
class excel:
	def __init__(self,fileUri):
		if os.path.exists(fileUri):
			self.uri = fileUri
		else:
			raise RuntimeError('无法找到该文件: %s' % fileUri)
	# excel 写入
	def excelWrite(self,data,title=[]):
		# 判断是否有标题存在
		# 判断是否有数据
		if len(data)>0:
			book = xlwt.Workbook(encoding='utf-8', style_compression=0)
			for key in sorted(data):
				row = 0
				col = 0
				sheet = book.add_sheet(key, cell_overwrite_ok=True)
				if len(title) > 0:
					for value in title:
						sheet .write(row,col,value)
						col += 1
					row += 1
				for line in data[key]:
					col = 0
					for value in line:
						sheet .write(row,col,value)
						col += 1
					row += 1
			book.save(self.uri)	
	
class excel2007:
	def __init__(self,fileUri):
		if not os.path.exists(fileUri):
			self.uri = fileUri
			self.file = openpyxl.Workbook()
			self.sheetno = 0
			self.sheets = []
		else:
			raise RuntimeError('文件已存在: %s' % fileUri)
	
	def addSheet(self,sheetName):
		# 判断工作表名称是否已存在
		if sheetName in self.sheets:
			raise RuntimeError('工作表名称已存在: %s' % sheetName)
		self.sheet = self.file.create_sheet(sheetName,self.sheetno)
		self.sheets.append(sheetName)
		self.sheetno += 1
	
	def writeData(self,data):
		for line in data:
			self.sheet.append(line)
	def save(self):	
		self.file.save(self.uri)

# excel 阅读类
# 只负责读取工作
'''
class excelread:
	def __init__(self,uri):
		if os.path.exists(self.uri):
			self.file = xlrd.open_workbook(self.uri)
		else:
			raise RuntimeError('无法找到该文件: %s' % fileUri)
	
	# 返回所有工作表的名称
	def getAllSheets(self):
		return self.file.sheet_names()
	
	# 根据表格名称或编号获取表格，优先根据名称，
	# 当名称为空时再使用编号进行获取，编号默认是第一个表格的编号
	def setTable(self,tableName='',tableNo=0):
		if not tableName == '':
			table = self.file.sheet_by_name(tableName)
		else:
			table = sheet_by_index(tableNo)
		self.table = table
	def getMaxRow(self):
		return self.table.nrows
	def getData(self):
		data = []
		for i in range(0,getMaxRow()):
			line = table.row_values(i)
			data.append(line)
		return data
'''
##################################################################################################################################		
