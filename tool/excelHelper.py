import xlwt,xlrd # excel 写入
import openpyxl # excel 2007
#import xlsxwriter # excel 2007解决方案性能优于openpyxl，功能只限于写入
import os

class excelHelper:
	# 三种模式 r w m (读取,写入,修改)
	def __init__(self,uri,type):
		self.uri = uri
		self.__checkFile()
		self.type = type
		if type=='r':
			self.file = excelReader(uri)
			#return excelReader(uri)
		elif type == 'w':
			self.file = excelWriter(uri,self.fileType)
			#return excelWriter(uri,self.fileType)
		elif type == 'm':
			self.workbook = map()
			self.file = excelReader(uri)
	# 检查文件
	def __checkFile(self):
		#检查文件类型
		if '.xlsx' in self.uri:
			self.fileType = 'excel2007'
		elif '.xls' in self.uri:
			self.fileType = 'excel2003'
		else:
			raise RuntimeError("无法识别文件类型:%s" % (self.uri))
		#检查文件是否存在
		self.fileActive = os.path.exists(self.uri)
	def copySheet(self,sheetName,newSheetName=''):
		self.file.setTable(sheetName)
		sheetData = self.file.getData()
		if newSheetName == '':
			newSheetName = sheetName
		self.workbook[newSheetName] = sheetData
	def copyWorkbook(self):
		sheets = self.file.getAllSheets()
		for sheetname in sheets:
			self.copySheet(sheetname)
	def save(self,uri):
		excel = excelWriter(uri,self.fileType)
		for sheetname in self.workbook.keys():
			excel.addSheet(sheetname)
			excel.writeData(self.workbook[sheetname])
		excel.save(uri)
		self.file.close()
		excel.close()
		
# excel 读取使用xlrd完成
# excel 阅读类
# 只负责读取工作
class excelReader:
	def __init__(self,uri):
		if os.path.exists(uri):
			self.file = xlrd.open_workbook(uri)
		else:
			raise RuntimeError('无法找到该文件: %s' % fileUri)
	
	# 返回所有工作表的名称
	def getAllSheets(self):
		return self.file.sheet_names()
	
	# 根据表格名称或编号获取表格，优先根据名称，
	# 当名称为空时再使用编号进行获取，编号默认是第一个表格的编号
	def setTable(self,tableName='',tableNo=0):
		if not tableName == '':
			table = self.file.sheet_by_name(tableName)
		else:
			table = self.file.sheet_by_index(tableNo)
		self.table = table
	def getMaxRow(self):
		return self.table.nrows
	def getData(self):
		data = []
		for i in range(0,self.getMaxRow()):
			line = self.table.row_values(i)
			data.append(line)
		return data
	def close(self):
		self.file.close()

class excelWriter:
	def __init__(self,fileUri,type):
		if not os.path.exists(fileUri):
			self.uri = fileUri
			if type == 'excel2003':
				self.file = xlwt.Workbook(encoding = 'utf-8')
			elif type == 'excel2007':
				self.file = openpyxl.Workbook()
				#self.file = self.file.active
			self.sheetno = 0
			self.sheets = []
		else:
			raise RuntimeError('文件已存在: %s' % fileUri)
	
	def addSheet(self,sheetName):
		# 判断工作表名称是否已存在
		if sheetName in self.sheets:
			raise RuntimeError('工作表名称已存在: %s' % sheetName)
		if type == 'excel2003':
			self.sheet = self.file.add_sheet(sheetName)
		elif type == 'excel2007':
			self.sheet = self.file.create_sheet(sheetName)
			#self.sheet.title = sheetName
		self.sheets.append(sheetName)
		self.sheetno += 1
	def __writeData2003(self,data):
		rowSize = len(data)-1
		for line in range(0,rowSize):
			colSize = len(line)-1
			for i in range(0,size):
				self.sheet.write(rowSize,colSize,data[rowSize][colSize])
	def __writeData2007(self,data):
		for line in data:
			self.sheet.append(line)
	def writeData(self,data):
		if type == 'excel2003':
			self.__writeData2003(data)
		elif type == 'excel2007':
			self.__writeData2007(data)
	def save(self):	
		self.file.save(self.uri)	

# 使用excelWriter实现		
class excel2007:
	def __init__(self,fileUri):
		if not os.path.exists(fileUri):
			self.uri = fileUri
			self.file = xlsxwriter.Workbook(fileUri)
			self.sheetno = 0
			self.sheets = []
		else:
			raise RuntimeError('文件已存在: %s' % fileUri)
	